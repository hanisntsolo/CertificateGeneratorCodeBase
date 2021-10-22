import cv2 as cv
import openpyxl

# template1.png is the template
# certificate
template_path = 'template12.jpeg'

# Excel file containing names of
# the participants
details_path = 'gsocOrgsList.xlsx'

# Output Paths
output_path = '/Users/dhsingh_old/Desktop/CertificateGenerator/Certificates'

# Setting the font size and font
# colour
font_size = 1
font_color = (0, 0, 0)

# Coordinates on the certificate where
# will be printing the name (set
# according to your own template)
coordinate_y_adjustment = 15
coordinate_x_adjustment = 20

# school name
# Coordinates on the certificate where
# will be printing the school name (set
# according to your own template)
coordinate_y_adjustment_school = -50
coordinate_x_adjustment_school = -150
# event name
# Coordinates on the certificate where
# will be printing the school name (set
# according to your own template)
coordinate_y_adjustment_event = -120
coordinate_x_adjustment_event = -50
# event name
# Coordinates on the certificate where
# will be printing the school name (set
# according to your own template)
coordinate_y_adjustment_year = -120
coordinate_x_adjustment_year = 500
# loading the details.xlsx workbook
# and grabbing the active sheet
obj = openpyxl.load_workbook(details_path)
sheet = obj.active

# printing for the first 10 names in the
# excel sheet
for i in range(2, 29):
    # grabs the row=i and column=1 cell
    # that contains the name value of that
    # cell is stored in the variable certi_name
    get_name = sheet.cell(row=i, column=2)
    certi_name = get_name.value

    get_school_name = sheet.cell(row=i, column=4)
    certi_school_name = get_school_name.value

    get_event_name = sheet.cell(row=i, column=5)
    certi_event_name = get_event_name.value

    # convergence year
    certi_year_name = "2021"
    # read the certificate template
    img = cv.imread(template_path)

    # choose the font from opencv
    # font = cv.FONT_HERSHEY_TRIPLEX
    font = cv.FONT_HERSHEY_COMPLEX

    thickness = 2
    # get the size of the name to be
    # printed
    text_size = cv.getTextSize(certi_name, font, font_size, 0)[0]

    text_size_school = cv.getTextSize(certi_school_name, font, font_size, 0)[0]

    text_size_event = cv.getTextSize(certi_event_name, font, font_size, 0)[0]

    text_size_year = cv.getTextSize(certi_year_name, font, font_size, 0)[0]

    # get the (x,y) coordinates where the
    # name is to written on the template
    # The function cv.putText accepts only
    # integer arguments so convert it into 'int'.
    text_x = (img.shape[1] - text_size[0]) / 2 + coordinate_x_adjustment
    text_y = (img.shape[0] + text_size[1]) / 2 - coordinate_y_adjustment
    text_x = int(text_x)
    text_y = int(text_y)
    cv.putText(img, certi_name,
               (text_x, text_y),
               font,
               font_size,
               font_color, thickness)
    # school name
    # get the (x,y) coordinates where the
    # name is to written on the template
    # The function cv.putText accepts only
    # integer arguments so convert it into 'int'.
    text_x_school = (img.shape[1] - text_size_school[0]) / 2 + coordinate_x_adjustment_school
    text_y_school = (img.shape[0] + text_size_school[1]) / 2 - coordinate_y_adjustment_school
    text_x_school = int(text_x_school)
    text_y_school = int(text_y_school)
    cv.putText(img, certi_school_name,
               (text_x_school, text_y_school),
               font,
               font_size,
               font_color, thickness)
    # event name
    # get the (x,y) coordinates where the
    # name is to written on the template
    # The function cv.putText accepts only
    # integer arguments so convert it into 'int'.
    text_x_event = (img.shape[1] - text_size_event[0]) / 2 + coordinate_x_adjustment_event
    text_y_event = (img.shape[0] + text_size_event[1]) / 2 - coordinate_y_adjustment_event
    text_x_event = int(text_x_event)
    text_y_event = int(text_y_event)
    cv.putText(img, certi_event_name,
               (text_x_event, text_y_event),
               font,
               font_size,
               font_color, thickness)
    # Output path along with the name of the
    # certificate generated

    # year
    # get the (x,y) coordinates where the
    # name is to written on the template
    # The function cv.putText accepts only
    # integer arguments so convert it into 'int'.
    text_x_year = (img.shape[1] - text_size_year[0]) / 2 + coordinate_x_adjustment_year
    text_y_year = (img.shape[0] + text_size_year[1]) / 2 - coordinate_y_adjustment_year
    text_x_year = int(text_x_year)
    text_y_year = int(text_y_year)
    cv.putText(img, certi_year_name,
               (text_x_year, text_y_year),
               font,
               font_size,
               font_color, thickness)
    certi_path = output_path + '/' + certi_name + certi_school_name + certi_event_name + '.png'

    # Save the certificate
    cv.imwrite(certi_path, img)